# User Task Manager (to-do-list)
# this programe that should allow users to 
# 1. add
# 2. view
# 3. update
# 4. delete
# tasks.
# all the data should be stored in a file.(accessible)

import openpyxl


class TasksManager:


    # class variable
    # dictionary to store all the tasks
    _all_tasks = {}


    # object intialization
    def __init__(self, username, password):
       self.username = username
       self.password = password
    

    # this function will add a task to a file
    def add_task(self, title, description, due_date):
        self._all_tasks['title'] = title
        self._all_tasks['description'] = description
        self._all_tasks['due_date'] = due_date

        database_wb = openpyxl.load_workbook('C:/Users/user/Desktop/Python_Learning/projectOne/database.xlsx')
        database_ws = database_wb['to_do_list']


        data_list = []
        for key in self._all_tasks:
            data_list.append(self._all_tasks[key])

        database_ws.append(data_list)
    
        database_wb.save('C:/Users/user/Desktop/Python_Learning/projectOne/database.xlsx')

        
    def view_tasks(self):
        database_wb = openpyxl.load_workbook('C:/Users/user/Desktop/Python_Learning/projectOne/database.xlsx')
        database_ws = database_wb['to_do_list']
        
        rows = database_ws.iter_rows(min_row=1, max_row=7, min_col=1, max_col=3)

        for row1, row2, row3 in rows:
            print(f'{row1.value}, {row2.value}, {row3.value}')
            print()
            

    def update_task(self, title):
        database_wb = openpyxl.load_workbook('C:/Users/user/Desktop/Python_Learning/projectOne/database.xlsx')
        database_ws = database_wb['to_do_list']

        _row = 1

        while True:
            if database_ws.cell(row=_row, column=1).value == title:
                new_title = input('New title: ')
                new_desctiption = input('New description: ')
                new_due_date = input('New due date (dd/mm/yr):')

                database_ws.cell(row=_row, column=1).value = new_title
                database_ws.cell(row=_row, column=2).value = new_desctiption
                database_ws.cell(row=_row, column=3).value = new_due_date     
                break
            else:
                _row += 1
                continue
        
        database_wb.save('C:/Users/user/Desktop/Python_Learning/projectOne/database.xlsx')


    def delete_task(self, title):
        database_wb = openpyxl.load_workbook('C:/Users/user/Desktop/Python_Learning/projectOne/database.xlsx')
        database_ws = database_wb['to_do_list']

        _row = 1

        while True:
            if database_ws.cell(row=_row, column=1).value == title:
                database_ws.delete_rows(_row, 1)
                break
            else:
                _row += 1
                continue


        database_wb.save('C:/Users/user/Desktop/Python_Learning/projectOne/database.xlsx')

